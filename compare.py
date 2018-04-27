#! /usr/bin/env python3

import openpyxl, sheetData_original, sheetData_client, \
    sheetData_client_map, sheetData_original_map, datetime
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.styles import colors

# subprocess.Popen(['/usr/local/bin/python3.6', 'dataExtracting.py'])

original = sheetData_original.allData
client = sheetData_client.allData

original_map = sheetData_original_map.mapping
client_map = sheetData_client_map.mapping

unmatched = {}

fileUnmatched = open('report_txt/改动过的.txt', 'w')
fileMissing = open('report_txt/缺的.txt', 'w')
fileAddition = open('report_txt/多的.txt', 'w')
fileUnmatched.write('----改动过的----\n')
fileMissing.write('----缺的----\n')
fileAddition.write('----多的----\n')

missing = set(original.keys()) - set(client.keys())
addition = set(client.keys()) - set(original.keys())
common = set(original.keys()).intersection(set(client.keys()))

print('Recording unmatched items...')
for i in common:
    if original[i] != client[i]:
        unmatched.update({i: list(set(client[i].items()).difference(set(original[i].items())))})
        fileUnmatched.write(str(i) + ': ' + str(set(set(client[i].items()).difference(set(original[i].items())))) + '\n')

print('Rocording missing items...')
if missing:
    for i in missing:
        fileMissing.write(str(i) + ': ' + str(original[i]['name']) + '\n')
else:
    fileMissing.write('>empty<')

print('Recording addition items...')
if addition:
    for i in addition:
        try:
            fileAddition.write(str(i) + ': ' + str(original[i]['name']) + '\n')
        except KeyError:
            continue
else:
    fileAddition.write('>empty<')

fileUnmatched.close()
fileAddition.close()
fileMissing.close()

# TODO: mark unmatched and addition item in client excel, store addition celss in a new worksheet

wb = openpyxl.load_workbook('copy/client_copy.xlsx')
ori_wb = openpyxl.load_workbook('copy/original_copy.xlsx')
ori_ws = ori_wb.active

sheet = wb['Sheet1']
sheet.freeze_panes = 'L6'
wb.create_sheet(title='缺的')
missingSheet = wb['缺的']
missingSheet.freeze_panes = 'K2'
missingSheet.column_dimensions['J'].width = 15

missingSheet['a1'] = '编号'
missingSheet['b1'] = '序号'
missingSheet['c1'] = '名称'
missingSheet['d1'] = '型号及规格'
missingSheet['e1'] = '单位'
missingSheet['f1'] = '数量'
missingSheet['g1'] = '送货单价'
missingSheet['h1'] = '金额'
missingSheet['i1'] = '送货单号'
missingSheet['j1'] = '送货日期'

row = 2
for i in missing:
    index = original_map[i]
    missingSheet['a' + str(row)] = ori_ws['a' + str(index)].value
    missingSheet['b' + str(row)] = ori_ws['b' + str(index)].value
    missingSheet['c' + str(row)] = ori_ws['c' + str(index)].value
    missingSheet['d' + str(row)] = ori_ws['d' + str(index)].value
    missingSheet['e' + str(row)] = ori_ws['e' + str(index)].value
    missingSheet['f' + str(row)] = ori_ws['f' + str(index)].value
    missingSheet['g' + str(row)] = ori_ws['g' + str(index)].value
    missingSheet['h' + str(row)] = '=f' + str(row) + '*g' + str(row)
    missingSheet['i' + str(row)] = ori_ws['i' + str(index)].value
    missingSheet['j' + str(row)] = ori_ws['j' + str(index)].value.date()
    row += 1

redFill = PatternFill(patternType='gray125', bgColor=colors.Color('f25c5c'), end_color=colors.Color('f25c5c'))
blueFill = PatternFill(patternType='gray125', bgColor=colors.Color('89b1e5'), end_color=colors.Color('89b1e5'))


for m in unmatched.keys():
    i = client_map[m]
    if unmatched[m][0][0] == 'serialNum':
        sheet['a' + str(i)].fill = redFill
    if unmatched[m][0][0] == 'sequenceNum':
        sheet['b' + str(i)].fill = redFill
    if unmatched[m][0][0] == 'name':
        sheet['c' + str(i)].fill = redFill
    if unmatched[m][0][0] == 'unit':
        sheet['e' + str(i)].fill = redFill
    if unmatched[m][0][0] == 'amount':
        sheet['f' + str(i)].fill = redFill
    if unmatched[m][0][0] == 'unitPrice':
        sheet['g' + str(i)].fill = redFill
    if unmatched[m][0][0] == 'deliverNum':
        sheet['i' + str(i)].fill = redFill
    if unmatched[m][0][0] == 'deliverDate':
        sheet['j' + str(i)].fill = redFill

for n in addition:
    index = client_map[n]
    sheet['D' + str(index)].fill = blueFill

wb.save('对比报告.xlsx')


