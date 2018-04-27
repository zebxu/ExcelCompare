#! /usr/bin/env python3

import openpyxl, pprint, threading
from openpyxl import Workbook
import os
from stat import S_IREAD, S_IRGRP, S_IROTH

# Load workbooks from excel files

print('Opening workbook')
wb_original = openpyxl.load_workbook('/Users/zeyangxu/Documents/Excel Automate Test/对的总对账单.xlsx')
wb_client = openpyxl.load_workbook('/Users/zeyangxu/Documents/Excel Automate Test/不对的.xlsx')

# Back up

wb_original.save('copy/original_copy.xlsx')
wb_client.save('copy/client_copy.xlsx')

# Get worksheet from workbook

sheet_original = wb_original.active
sheet_client = wb_client.active

# Function for retrieving data from the worksheet to data
def readRows(sheet, fileName):
    allData ={}
    mapping = {}
    print('Reading rows...')
    for row in range(5, sheet.max_row + 1):
        print('Collect data from row %s' % row)
        # Each row in the spreadsheet has data for one census tract.
        serialNum = sheet['A' + str(row)].value
        sequenceNum = sheet['B' + str(row)].value
        name = sheet['C' + str(row)].value
        model = sheet['D' + str(row)].value
        unit = sheet['E' + str(row)].value
        amount = sheet['F' + str(row)].value
        unitPrice = sheet['G' + str(row)].value
        # totalPrice = sheet['H' + str(row)].value
        deliverNum = sheet['I' + str(row)].value
        deliverDate = sheet['J' + str(row)].value

        # setdefault do nothing when the key already exist

        allData.setdefault((model, sequenceNum), {'name': name,
                                   'serialNum': serialNum,
                                   'sequenceNum': sequenceNum,
                                   'unit': unit,
                                   'amount': amount,
                                   'unitPrice': unitPrice,
                                   'deliverNum': deliverNum,
                                   'deliverDate': deliverDate})
        mapping.update({(model, sequenceNum): row})

    print('Writing results...')
    resultFile = open(fileName + '.py', 'w')
    mapFile = open(fileName + '_map.py', 'w')
    resultFile.write('import datetime\n')
    resultFile.write('allData = ' + pprint.pformat(allData))
    mapFile.write('mapping = ' + pprint.pformat(mapping))
    resultFile.close()
    mapFile.close()
    # Change file to read-only
    # os.chmod(fileName + '.py', S_IREAD | S_IRGRP | S_IROTH)
    print('Done')
    return


# Second thread
threadObj = threading.Thread(target=readRows, args=[sheet_original, 'sheetData_original'])
threadObj.start()

# readRows(sheet_original, 'sheetData_original')
readRows(sheet_client, 'sheetData_client')

print('Main thread end.')
